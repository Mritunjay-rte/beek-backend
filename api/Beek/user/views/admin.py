from django.contrib.auth.tokens import PasswordResetTokenGenerator
from rest_framework.generics import ListAPIView
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.generics import GenericAPIView
from rest_framework.views import APIView
from user.models import User, PasswordReset, UserSubscriptions
from user.serializers import UserSerializer, UserAccountSerializer, UserReportListSerializer, UserAPIUsageReportListSerializer, UserAPIUsageDetailsListSerializer
from health.models import ExternalAPILog, ServiceProvider
from user.filters import UserFilter, UserReportFilter, UserAPIUsageReportFilter
from beek.responses import SuccessResponse, ErrorResponse
from user.views.user import send_password_reset_email
from django.db.models import F, Case, CharField, OuterRef, Subquery, Value, When
from django.db.models.functions import Lower
from drf_yasg.utils import swagger_auto_schema
from user.utils import get_user_report_queryset, get_user_api_usage_report_queryset
from drf_yasg import openapi
from rest_framework import status
from django.db.models import OuterRef, Subquery
from django.utils.timezone import now, utc
from datetime import timedelta
from datetime import datetime
import csv
import openpyxl
import io
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill
import stripe


class UsersListView(ListAPIView):
    """
        API view to retrieve a list of all users for admin access.

        This view allows an authenticated admin user to retrieve a list of all users in the system.
        It provides a paginated list of users, including their details such as name, email, and account status.
        This view is restricted to admin users only, ensuring that only authorized personnel can access the user data.

        Permissions:
            - Admin users only. Non-admin users will not have access to this endpoint.

        Request:
            - No specific parameters are required for listing users.
            - Pagination options can are implemented to retrieve users in pages.

        Response:
            - A list of user objects, each containing details such as name, email, and account status.
            - Pagination details are included in the response.

        Returns:
            - A paginated list of user data for all users in the system.
    """
    queryset = User.objects.filter(groups__name='USER', is_active=True)
    serializer_class = UserSerializer
    filter_backends = (DjangoFilterBackend,
                       filters.SearchFilter)
    filterset_class = UserFilter
    search_fields = ['name', 'email']
    permission_classes = [IsAdminUser]


    def get_queryset(self):
        """
            Returns the queryset for retrieving a list of users, ordered by specified criteria.

            This method retrieves all users from the database and applies ordering to the queryset. 
            The ordering can be based on fields such as 'name', 'email', 'created_at', etc. 
            The default ordering can be specified here, but custom ordering can be applied by the client.

            If no ordering criteria are provided, the queryset will be ordered by the default field'name'.

            Returns:
                QuerySet: A queryset of user objects, ordered according to the specified ordering.
        """
        ordering = self.request.query_params.get('ordering', None)
        if ordering:
            if ordering.startswith('-'):
                field_name = ordering[1:]
                field_type = self.queryset.model._meta.get_field(field_name).get_internal_type()
                if field_type == 'DateTimeField':
                    queryset = self.queryset.order_by(F(field_name).desc())
                else:
                    queryset = self.queryset.order_by(Lower(F(field_name)).desc())
            else:
                field_name = ordering
                field_type = self.queryset.model._meta.get_field(field_name).get_internal_type()
                if field_type == 'DateTimeField':
                    queryset = self.queryset.order_by(F(field_name))
                else:
                    queryset = self.queryset.order_by(Lower(F(field_name)))
        else:
            queryset = self.queryset
        
        return queryset
    
    @swagger_auto_schema(
        operation_summary="List all users in admin panel",
        operation_description="""
            This API lists all the users.
        """,
        tags=["user"],
        manual_parameters=[],
        operation_id='listUsers',
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class RestrictUserAccountView(GenericAPIView):
    """
    Restricts or deactivates a user account based on the provided request data.

    This method handles the logic for restricting or deactivating a user account. It expects the 
    `user_id` in the request body. The method first validates the input, 
    checks if the user exists, and then either deactivates/activates the user's account.

    Args:
        request (Request): The HTTP request containing the `user_id`.

    Returns:
        Response: A JSON response indicating the success or failure of the account restriction operation. 
                  The response includes a message and may also return relevant data depending on the action taken.
                  
    Raises:
        ValidationError: If the provided data is invalid or missing required fields (e.g., `user_id`).
        Exception: In case of unexpected errors during the process.
    """
    permission_classes = [IsAuthenticated]
    serializer_class = UserAccountSerializer

    @swagger_auto_schema(
        operation_summary="Restrict User Account",
        operation_description="""
            This API allows administrators or system processes to restrict or deactivate a user's account.
            It requires the user ID of the account to be restricted.
        """,
        operation_id='restrictAccount',
        tags=["user"],
    )
    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.user

        user.is_blocked = not user.is_blocked
        user.save()
        if user.is_blocked:
            user_subscriptions_obj = UserSubscriptions.objects.filter(
                user_id=user.id, is_active=True, is_expired=False).order_by('-created_at').first()
            if user_subscriptions_obj:
                stripe.Subscription.modify(
                    user_subscriptions_obj.stripe_id,
                    cancel_at_period_end=True
                )
                user_subscriptions_obj.payment_method_deleted = True
                user_subscriptions_obj.cancelled_on = datetime.now(utc)
                user_subscriptions_obj.save()
            return SuccessResponse(message="User account successfully restricted.")
        else:
            return SuccessResponse(message="User account successfully reactivated.")


class AdminPasswordResetView(GenericAPIView):
    """
        API View to initiate a password reset for a user on behalf of an administrator.

        This view allows an admin user to request a password reset link for another user. 
        The admin must provide the user's uuid in the request, and if the user 
        exists and meets the criteria (e.g., active and not blocked), a password reset email 
        will be sent to the user.

        Attributes:
            serializer_class (Serializer): Serializer for validating the input.
            permission_classes (list): Permissions required to access this view.

        Methods:
            post(request):
                Initiates the password reset process for the specified user.
    """
    permission_classes = [IsAuthenticated]
    serializer_class = UserAccountSerializer

    @swagger_auto_schema(
        operation_summary="Admin Request Password Reset Link",
        operation_description="""
            This API allows an admin to request a password reset link for a specific user. 
            The admin must provide the email address of the user whose password needs to be reset. 
            If the email is associated with an account, a password reset link will be sent to that user's email address.
        """,
        operation_id='adminRequestPasswordResetLink',
        tags=["user"]
    )
    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.user

        if user:
            token_generator = PasswordResetTokenGenerator()
            token = token_generator.make_token(user)
            reset = PasswordReset(user=user, token=token)
            reset.save()
            send_password_reset_email(user, token)
        return SuccessResponse(message='A password reset link has been sent to the user')
    

class UserReportListView(ListAPIView):
    """
        API View to retrieve a list of user activity reports for administrative users.

        This view provides detailed information about users, including their activity, 
        status, subscription etc. It is accessible only to admin users and 
        supports filtering, ordering, and pagination for efficient data retrieval.

        Attributes:
            serializer_class (Serializer): Serializer used to format the output data.
            permission_classes (list): Permissions required to access this view. 
                Typically, only admin users are allowed.
            queryset (QuerySet): The default queryset to retrieve user reports.
            filter_backends (list): Filter and ordering backends for querying the data.
            pagination_class (Pagination): Pagination class used to limit the results per page.

        Methods:
            get_queryset():
                Returns a filtered and ordered queryset of user reports based on 
                request parameters.
    """
    queryset = User.objects.filter(groups__name='USER', is_active=True).order_by('name')
    serializer_class = UserReportListSerializer
    filter_backends = (DjangoFilterBackend,
                       filters.SearchFilter)
    filterset_class = UserReportFilter
    search_fields = ['name', 'email', 'user_id']
    permission_classes = [IsAdminUser]


    def get_queryset(self):
        ordering = self.request.query_params.get('ordering')
        return get_user_report_queryset(self.queryset, ordering)
    
    @swagger_auto_schema(
        operation_summary="List all users report in admin panel",
        operation_description="""
            This API lists all the users report.
        """,
        tags=["user"],
        manual_parameters=[
            openapi.Parameter(
                name="created_at_after",
                in_=openapi.IN_QUERY,
                description="Filter users created on or after this date (format: YYYY-MM-DD)",
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE
            ),
            openapi.Parameter(
                name="created_at_before",
                in_=openapi.IN_QUERY,
                description="Filter users created on or before this date (format: YYYY-MM-DD)",
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE
            ),
            openapi.Parameter(
                name="last_login_at_after",
                in_=openapi.IN_QUERY,
                description="Filter users who logged in on or after this date (format: YYYY-MM-DD)",
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE
            ),
            openapi.Parameter(
                name="last_login_at_before",
                in_=openapi.IN_QUERY,
                description="Filter users who logged in on or before this date (format: YYYY-MM-DD)",
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE
            ),
            openapi.Parameter(
                name="billing_period_start_after",
                in_=openapi.IN_QUERY,
                description="Filter users with last payment on or after this date (format: YYYY-MM-DD)",
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE
            ),
            openapi.Parameter(
                name="billing_period_start_before",
                in_=openapi.IN_QUERY,
                description="Filter users with last payment on or before this date (format: YYYY-MM-DD)",
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE
            ),
            openapi.Parameter(
                name="subscription_status",
                in_=openapi.IN_QUERY,
                description="Filter users by subscription status (e.g., Active, Expired, Trial, etc.)",
                type=openapi.TYPE_STRING
            ),
            openapi.Parameter(
                name="ordering",
                in_=openapi.IN_QUERY,
                description="Fields to sort by (e.g., user_id, name, etc.)",
                type=openapi.TYPE_STRING
            ),
        ],
        operation_id='listUsersReport',
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class UserReportMetrics(APIView):
    """
        API View to retrieve user metrics for administrative users.

        This view provides key metrics related to user subscriptions, including:
        - The total number users.
        - The number of active subscriptions.
        - The number of expired subscriptions.
        - The number of users in trial period.
        - New signups
        - Upcoming renewals

        It is designed to give administrators an overview of the current state of user subscriptions and 
        their changes over time.

        Attributes:
            permission_classes (list): Permissions required to access this view. 
                Typically restricted to admin users.

        Methods:
            get(request):
                Handles GET requests to retrieve subscription metrics and relevant data.
    """
    permission_classes = [IsAdminUser]

    @swagger_auto_schema(
        operation_summary="User report metrics for admin",
        operation_description="""
            This API allows an admin to view user report metrics.
            This includes total users, active subscriptions, expired subscriptions, trial period, new signups and upcoming renewals
        """,
        operation_id='userReportMetrics',
        tags=["user"]
    )
    def get(self, request, *args, **kwargs):
        try:
            obj_user = User.objects.filter(groups__name='USER', is_active=True)
            data = {}
            if obj_user.exists():
                
                # Total user count
                data['total_users'] = obj_user.count()

                latest_sub_query = UserSubscriptions.objects.filter(user_id=OuterRef('pk')).order_by('-created_at')
                annotated_queryset = obj_user.annotate(
                    latest_status=Subquery(latest_sub_query.values('status')[:1]),
                    subscription_status=Case(
                        When(deleted_at__isnull=False, then=Value('Terminated')),
                        When(latest_status='trial', then=Value('Trial')),
                        When(latest_status='expired', then=Value('Expired')),
                        When(latest_status__in=['active', 'cancelled'], then=Value('Active')),
                        default=Value('Pending'),
                        output_field=CharField()
                    )
                )
                
                # Active users count
                data['active_subscriptions'] = annotated_queryset.filter(
                    subscription_status__in=["Active", "Cancelled"]
                ).distinct().count()

                # Expired user subscriptions
                data['expired_subscriptions'] = annotated_queryset.filter(
                    subscription_status="Expired"
                ).distinct().count()

                # Users in trial period
                data['trial_period'] = annotated_queryset.filter(
                    subscription_status="Trial"
                ).distinct().count()

                # New signups in last 30 days
                thirty_days_ago = datetime.now() - timedelta(days=30)
                data['new_signups'] = obj_user.filter(created_at__date__gte=thirty_days_ago).count()

                # Upcoming renewals in 7 days
                week = datetime.now() + timedelta(days=6)
                data['upcoming_renewals'] = annotated_queryset.filter(
                    subscriptions__billing_period_end__date__gte=datetime.now(),
                    subscriptions__billing_period_end__date__lte=week,
                    subscriptions__is_expired=False,
                    subscription_status='Active'
                ).distinct().count()

            else:
                data = {
                    'total_users': 0,
                    'active_subscriptions': 0,
                    'expired_subscriptions': 0,
                    'trial_period': 0,
                    'new_signups': 0,
                    'upcoming_renewals': 0
                }
            return SuccessResponse(message="User report metrics", data=data)

        except Exception as e:
            print(str(e))
            return ErrorResponse(message="Something went wrong", status_code=status.HTTP_400_BAD_REQUEST,errors=str(e))


class ExportUserReportView(GenericAPIView):
    """
    API View to export user reports as downloadable files.

    This view allows administrators to export user data, including metrics 
    related to user activity, subscriptions, and other relevant information, 
    in a structured file format such as .csv or .xlsx.

    Attributes:
        permission_classes (list): Specifies the permissions required to access this view.
                                   Typically restricted to administrative users.
        serializer_class (Serializer): Defines the serializer used for validating and processing input data.

    Params:
        doc_type : mandatory...either csv or xlsx
    
    Methods:
            get(request)
    """
    queryset = User.objects.filter(groups__name='USER', is_active=True).order_by('name')
    permission_classes = [IsAdminUser]
    filter_backends = [DjangoFilterBackend]
    filterset_class = UserReportFilter
    serializer_class = UserReportListSerializer

    def get_queryset(self):
        """
        Annotate and filter the queryset as per UserReportListView.
        """
        ordering = self.request.query_params.get('ordering')
        return get_user_report_queryset(self.queryset, ordering)

    @swagger_auto_schema(
        operation_summary="Export User Report",
        operation_description="""
            Export user report in the requested format (CSV or XLSX).
        """,
        tags=["user"],
        manual_parameters=[
            openapi.Parameter(
                name="doc_type",
                in_=openapi.IN_QUERY,
                description="Format of the exported file ('csv' or 'xlsx').",
                type=openapi.TYPE_STRING,
                required=True,
                enum=['csv', 'xlsx']
            )
        ],
        operation_id='exportUsersReport',
    )
    def get(self, request, *args, **kwargs):
        format = request.query_params.get('doc_type', 'csv').lower()
        if format not in ['csv', 'xlsx']:
            return ErrorResponse(message="Invalid format. Use 'csv' or 'xlsx'.", status_code=status.HTTP_400_BAD_REQUEST,errors={})

        queryset = self.filter_queryset(self.get_queryset())
        data = queryset.values(
            'user_id', 'name', 'email', 'phone_number', 'created_at', 'subscription_status', 
            'last_login_at', 'billing_period_start', 'billing_period_end', 
        )

        if format == 'csv':
            return self.export_as_csv(data)
        elif format == 'xlsx':
            return self.export_as_xlsx(data)
        
    def get_file_name(self, extension):
        """
        Generate file name in the format 'UserReport_HH:MM:SS; Mon-DD-YYYY'.
        """
        timestamp = now().strftime('%H_%M_%S_%b-%d-%Y')
        return f"UserReport_{timestamp}.{extension}"

    def export_as_csv(self, data):
        """
        Export the data to CSV format.
        """
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(['#', 'User ID', 'Name', 'Email ID', 'Phone Number', 'Sign-Up Date', 'Subscription Status', 'Last Login',
                         'Last Payment', 'Next Renewal'])
        for index, item in enumerate(data, start=1):
            item['created_at'] = item['created_at'].strftime('%b-%d-%Y') if item['created_at'] else "-"
            item['billing_period_start'] = item['billing_period_start'].strftime('%b-%d-%Y') if item['billing_period_start'] else "-"
            item['billing_period_end'] = item['billing_period_end'].strftime('%b-%d-%Y') if item['billing_period_end'] else "-"
            item['last_login_at'] = item['last_login_at'].strftime("%H:%M:%S, %b-%d-%Y") if item['last_login_at'] else "-"
            if item['subscription_status'] == 'Terminated':
                item.update(email="-", phone_number="-")

            # Now write the formatted values to CSV
            writer.writerow([index, item['user_id'], item['name'], item['email'], item['phone_number'], 
                             item['created_at'], item['subscription_status'],item['last_login_at'],
                             item['billing_period_start'],item['billing_period_end'],
                            ])
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.get_file_name("csv")}"'
        return response
    
    def export_as_xlsx(self, data):
        """
        Export the data to XLSX format.
        """
        # Create a new workbook and set the title for the sheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'User Report'
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        # Write header
        headers = ['#', 'User ID', 'Name', 'Email ID', 'Phone Number', 'Sign-Up Date', 'Subscription Status', 'Last Login',
                'Last Payment', 'Next Renewal']
        worksheet.append(headers)
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        # Write data rows
        for index, item in enumerate(data, start=1): 
            # Format the date fields
            item['created_at'] = item['created_at'].strftime('%b-%d-%Y') if item.get('created_at') else "-"
            item['billing_period_start'] = item['billing_period_start'].strftime('%b-%d-%Y') if item.get('billing_period_start') else "-"
            item['billing_period_end'] = item['billing_period_end'].strftime('%b-%d-%Y') if item.get('billing_period_end') else "-"
            item['last_login_at'] = item['last_login_at'].strftime("%H:%M:%S, %b-%d-%Y") if item.get('last_login_at') else "-"
            if item['subscription_status'] == 'Terminated':
                item.update(email="-", phone_number="-")

            # Append the data to the worksheet
            worksheet.append([
                index,
                item.get('user_id'),
                item.get('name'),
                item.get('email'),
                item.get('phone_number'),
                item.get('created_at'),
                item.get('subscription_status'),
                item.get('last_login_at'),
                item.get('billing_period_start'),
                item.get('billing_period_end'),
            ])

        # Create a buffer to store the XLSX file
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)  # Go back to the start of the buffer to read it
        
        # Prepare the response with the buffer content as an XLSX file
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Set the content-disposition header to make the file downloadable
        response['Content-Disposition'] = f'attachment; filename="{self.get_file_name("xlsx")}"'
        return response


class UserAPIUsageReportListView(ListAPIView):
    """
        API View to list and display detailed user API usage reports.

        This view provides a paginated list of API usage statistics for each users, including:
        - Total number of API calls.
        - Number of initial and refresh API calls.
        - Last API call timestamp.
        - Last login timestamp.
        - User status (e.g., Active, Inactive).

        This is typically used by administrators to monitor API usage, detect patterns, and analyze user behavior.

        Attributes:
            serializer_class (Serializer): Specifies the serializer used to format API usage data for the response.
            permission_classes (list): Specifies the permissions required to access this view, typically restricted to admin users.
            queryset (QuerySet): The base queryset used to retrieve API usage data for users.
            filter_backends (list): Defines the filtering, searching, and ordering backends.
    """
    
    queryset = User.objects.filter(groups__name='USER', is_active=True)
    serializer_class = UserAPIUsageReportListSerializer
    filter_backends = (DjangoFilterBackend,
                       filters.SearchFilter)
    filterset_class = UserAPIUsageReportFilter
    search_fields = ['name', 'user_id']
    permission_classes = [IsAdminUser]


    def get_queryset(self):
        ordering = self.request.query_params.get('ordering')
        return get_user_api_usage_report_queryset(self.queryset, ordering)
    
    @swagger_auto_schema(
        operation_summary="List all users API usage report in admin panel",
        operation_description="""
            This API lists all the users API usage report.
        """,
        tags=["user"],
        manual_parameters=[
            openapi.Parameter(
                name="last_login_at_after",
                in_=openapi.IN_QUERY,
                description="Filter users who logged in on or after this date (format: YYYY-MM-DD)",
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE
            ),
            openapi.Parameter(
                name="last_login_at_before",
                in_=openapi.IN_QUERY,
                description="Filter users who logged in on or before this date (format: YYYY-MM-DD)",
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE
            ),
            openapi.Parameter(
                name="user_status",
                in_=openapi.IN_QUERY,
                description="Filter users by multiple statuses (e.g., Active, Inactive, Restricted, Deleted.)",
                type=openapi.TYPE_STRING
            ),
            openapi.Parameter(
                name="provider",
                in_=openapi.IN_QUERY,
                description="Filter users by multiple provider references (e.g., 4728, 4851, etc.)",
                type=openapi.TYPE_STRING
            ),
            openapi.Parameter(
                name="ordering",
                in_=openapi.IN_QUERY,
                description="Fields to sort by (e.g., user_id, name, etc.)",
                type=openapi.TYPE_STRING
            ),
        ],
        operation_id='listUsersAPIUsageReport',
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class ApiUsageReportMetrics(APIView):
    """
    API View to provide aggregated metrics for API usage.

    This view returns various summary metrics related to API usage across the system, 
    helping administrators monitor and analyze API consumption patterns. 

    Metrics provided may include:
    - Active API Users.
    - Total number of API calls.
    - Number of initial API calls.
    - Number of refresh API calls.
    - Average API calls per user.
    - Number of connected providers

    Attributes:
        permission_classes (list): Specifies the permissions required to access this view, typically restricted to admin users.
    Methods:
        get(self, request, *args, **kwargs):
            Handles GET requests and returns aggregated API usage metrics.

    Returns:
        Response: A JSON object containing various API usage metrics.
    """
    permission_classes = [IsAdminUser]

    @swagger_auto_schema(
        operation_summary="Api usage report metrics for admin",
        operation_description="""
            This API allows an admin to view 1UP api usage report metrics.
            This includes active api users, total api calls, initial api calls, refresh api calls, average api calls per user and connected providers
        """,
        operation_id='apiUsageReportMatrics',
        tags=["user"],
        manual_parameters=[
            openapi.Parameter(
                name="start_date",
                in_=openapi.IN_QUERY,
                description="Filter api usage metrics based on api's called on or after this date (format: YYYY-MM-DD)",
				type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE
            ),
            openapi.Parameter(
				name="end_date",
                in_=openapi.IN_QUERY,
                description="Filter api usage metrics based on api's called on or before this date (format: YYYY-MM-DD)",
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE
            ),
        ]
    )
    def get(self, request, *args, **kwargs):
        try:
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            data = {
                'active_api_users': 0,
                'total_api_calls': 0,
                'initial_api_calls': 0,
                'refresh_api_calls': 0,
                'average_api_calls': 0,
                'connected_providers': 0
            }
            obj_externalapi_log = ExternalAPILog.objects
            obj_serviceprovider = ServiceProvider.objects
            if obj_externalapi_log.exists():
                if start_date and end_date:
                    try:
                        start_date = datetime.strptime(start_date, "%Y-%m-%d")
                        end_date = datetime.strptime(end_date, "%Y-%m-%d")
                    except (TypeError, ValueError):
                        return ErrorResponse(message="Invalid date format. Use YYYY-MM-DD.", status_code=status.HTTP_400_BAD_REQUEST)
                    obj_externalapi_log = obj_externalapi_log.filter(requested_at__range=(start_date.strftime("%Y-%m-%d 00:00:00"), end_date.strftime("%Y-%m-%d 23:59:59")))
                    obj_serviceprovider = obj_serviceprovider.filter(created_at__range=(start_date.strftime("%Y-%m-%d 00:00:00"), end_date.strftime("%Y-%m-%d 23:59:59")))
                data['active_api_users'] = obj_externalapi_log.values('user').distinct().count()
                data['total_api_calls'] = obj_externalapi_log.count()
                data['initial_api_calls'] = obj_externalapi_log.filter(is_initial_sync=True).count()
                data['refresh_api_calls'] = obj_externalapi_log.filter(is_initial_sync=False).count()
                if data['active_api_users'] > 0:
                    data['average_api_calls'] = data['total_api_calls'] / data['active_api_users']
                data['connected_providers'] = obj_serviceprovider.values('provider_id').distinct().count()
            return SuccessResponse(message="Api usage report metrics", data=data)
        except Exception as e:
            print(str(e))
            return ErrorResponse(message="Something went wrong", status_code=status.HTTP_400_BAD_REQUEST,errors=str(e))


class ExportUserAPIUsageReportView(GenericAPIView):
    """
    API View to export user API usage reports in a downloadable file format.

    This view allows administrators to export a report containing detailed API usage statistics 
    for all users. The report is typically generated in CSV or Excel format and includes 
    metrics such as total API calls, initial and refresh API calls, last API call, last login, 
    and user status.

    Attributes:
        permission_classes (list): Specifies the permissions required to access this view, 
        typically restricted to admin users.

    Report Fields:
        - `#`: Auto-incremented row number.
        - `User ID`: Unique identifier of the user.
        - `Name`: Full name of the user.
        - `Connected Providers`: Number of third-party providers connected to the user account.
        - `Total API Calls`: Total number of API requests made by the user.
        - `Initial API Calls`: Number of initial API requests made by the user.
        - `Refresh API Calls`: Number of refresh API requests made by the user.
        - `Last API Call`: Timestamp of the last API request made.
        - `Last Login`: Timestamp of the user's last login.
        - `User Status`: Current status of the user (e.g., Active, Blocked, Deleted).

    Methods:
        get(self, request, *args, **kwargs):
            Handles POST requests and generates the API usage report file.
            Returns a downloadable file as the response.

    Returns:
        Response: A file download response containing the API usage report, or an error response if the export fails.
    """
    queryset = User.objects.filter(groups__name='USER', is_active=True)
    serializer_class = UserAPIUsageReportListSerializer
    filter_backends = (DjangoFilterBackend,
                       filters.SearchFilter)
    filterset_class = UserAPIUsageReportFilter
    search_fields = ['name', 'user_id']
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        ordering = self.request.query_params.get('ordering')
        return get_user_api_usage_report_queryset(self.queryset, ordering)
    
    @swagger_auto_schema(
        operation_summary="Export User API Usage Report",
        operation_description="""
            Export user api usage report in the requested format (CSV or XLSX).
        """,
        tags=["user"],
        manual_parameters=[
            openapi.Parameter(
                name="doc_type",
                in_=openapi.IN_QUERY,
                description="Format of the exported file ('csv' or 'xlsx').",
                type=openapi.TYPE_STRING,
                required=True,
                enum=['csv', 'xlsx']
            )
        ],
        operation_id='exportAPIUsageReport',
    )
    def get(self, request, *args, **kwargs):
        format = request.query_params.get('doc_type', 'csv').lower()
        if format not in ['csv', 'xlsx']:
            return ErrorResponse(message="Invalid format. Use 'csv' or 'xlsx'.", status_code=status.HTTP_400_BAD_REQUEST,errors={})

        queryset = self.filter_queryset(self.get_queryset())
        data = queryset.values(
            'user_id', 'name', 'provider_count', 'total_api_calls', 'initial_api_calls', 'refresh_api_calls', 
            'last_api_call', 'last_login_at', 'user_status'
        )

        if format == 'csv':
            return self.export_as_csv(data)
        elif format == 'xlsx':
            return self.export_as_xlsx(data)
        
    def get_file_name(self, extension):
        """
        Generate file name in the format 'APIUsageReport_HH:MM:SS; Mon-DD-YYYY'.
        """
        timestamp = now().strftime('%H_%M_%S_%b_%d_%Y')
        return f"UsageReport_{timestamp}.{extension}"

    def export_as_csv(self, data):
        """
        Export the data to CSV format.
        """
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(['#', 'User ID', 'Name', 'Connected Providers', 'Total API Calls', 'Initial API Calls', 'Refresh API Calls', 'Last API Call',
                         'Last Login', 'User Status'])
        for index, item in enumerate(data, start=1):
            last_login_at = item.get('last_login_at').strftime('%H:%M:%S, %b-%d-%Y') if item.get('last_login_at') else "-"
            last_api_call = item.get('last_api_call').strftime('%H:%M:%S, %b-%d-%Y') if item.get('last_api_call') else "-"
            # Now write the formatted values to CSV
            writer.writerow([index, item['user_id'], item['name'], item['provider_count'], item['total_api_calls'], 
                             item['initial_api_calls'], item['refresh_api_calls'], last_api_call, last_login_at,
                             item['user_status'],
                            ])
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.get_file_name("csv")}"'
        return response
    
    def export_as_xlsx(self, data):
        """
        Export the data to XLSX format.
        """
        # Create a new workbook and set the title for the sheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'User API Usage Report'
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Write header
        headers = ['#', 'User ID', 'Name', 'Connected Providers', 'Total API Calls', 'Initial API Calls', 'Refresh API Calls', 'Last API Call',
                         'Last Login', 'User Status']
        worksheet.append(headers)

        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
        # Write data rows
        for index, item in enumerate(data, start=1): 
            # Append the data to the worksheet
            last_login_at = item.get('last_login_at').strftime('%H:%M:%S, %b-%d-%Y') if item.get('last_login_at') else "-"
            last_api_call = item.get('last_api_call').strftime('%H:%M:%S, %b-%d-%Y') if item.get('last_api_call') else "-"

            worksheet.append([
                index,
                item.get('user_id'),
                item.get('name'),
                item.get('provider_count'),
                item.get('total_api_calls'),
                item.get('initial_api_calls'),
                item.get('refresh_api_calls'),
                last_api_call,
                last_login_at,
                item.get('user_status'),
            ])

        # Create a buffer to store the XLSX file
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)  # Go back to the start of the buffer to read it
        
        # Prepare the response with the buffer content as an XLSX file
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Set the content-disposition header to make the file downloadable
        response['Content-Disposition'] = f'attachment; filename="{self.get_file_name("xlsx")}"'
        return response


class UserAPIUsageDetailsListView(ListAPIView):
    """
    Retrieve detailed API usage logs for a specific user based on the type of API usage.

    This API view allows the administrator to retrieve a detailed list of API usage logs for a particular user,
    including different types of API calls (total, initial, and refresh). The results are filtered based on the 
    `user_id` and `type` query parameters.

    Query Parameters:
        - `user_id` (required): The ID of the user for whom API usage details are being retrieved.
        - `type` (optional): The type of API logs to retrieve. Can be one of the following:
            - `'total'`: All API calls made by the user.
            - `'initial'`: API calls related to the user's initial interaction.
            - `'refresh'`: API calls related to refreshing the user's session.

    Response:
        Returns a paginated list of API usage records for the specified user and type, including relevant fields like
        the number of API calls, timestamps, and user status.

    Methods:
        get:
            Returns the filtered queryset based on the `user_id` and `type` query parameters.
    """

    queryset = ExternalAPILog.objects.all()
    serializer_class = UserAPIUsageDetailsListSerializer
    permission_classes = [IsAdminUser]
    
    def get_queryset(self):
        """
        filter the base queryset with query params such as user_id and api call type
        """
        user_id = self.request.query_params.get('user_id')
        log_type = self.request.query_params.get('type')

        # Validate `user_id` parameter
        if not user_id:
            return ExternalAPILog.objects.none()  # Return empty queryset if no user_id provided

        queryset = ExternalAPILog.objects.filter(user_id=user_id)

        # Filter based on `type` parameter
        if log_type == 'initial':
            queryset = queryset.filter(is_initial_sync=True)
        elif log_type == 'refresh':
            queryset = queryset.filter(is_initial_sync=False)
        elif log_type != 'total':
            return ExternalAPILog.objects.none()  # Return empty queryset for invalid type

        return queryset.order_by('-requested_at')
    
    @swagger_auto_schema(
        operation_summary="Retrieve API Usage Logs for a User",
        operation_description="""
            This API retrieves detailed API usage logs for a specific user. 
            You can filter the logs based on the type of API call (e.g., total, initial, refresh).
        """,
        tags=["user"],
        manual_parameters=[
            openapi.Parameter(
                name="user_id",
                in_=openapi.IN_QUERY,
                description="The ID of the user whose API logs are to be retrieved. This parameter is required.",
                type=openapi.TYPE_STRING,
                required=True
            ),
            openapi.Parameter(
                name="type",
                in_=openapi.IN_QUERY,
                description="""
                    The type of API logs to filter by:
                    - `total`: Retrieve all API logs for the user.
                    - `initial`: Retrieve only initial API synchronization logs.
                    - `refresh`: Retrieve only refresh API synchronization logs.
                    If no valid type is provided, the result will be empty.
                """,
                type=openapi.TYPE_STRING,
                required=False,
                enum=["total", "initial", "refresh"]
            ),
        ],
        operation_id='getUserAPIUsageLogs',
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)